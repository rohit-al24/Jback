"""
Administration API views.

Roles hierarchy:
  master_admin  → manages colleges, assigns college_admins
  college_admin → manages departments, sensei, classes, students, requests
  sensei        → views assigned classes + students
  staff         → views own mentees
  student       → views/sets own mentor, sends change requests
"""
from __future__ import annotations

import io
import json
import random
import string

import openpyxl
from django.contrib.auth import get_user_model
from django.contrib.auth.hashers import make_password
from django.db import IntegrityError, transaction
from django.http import HttpRequest, HttpResponse, JsonResponse

from core.decorators import api_login_required, require_role
from core.models import College
from .models import (
    ClassStudent,
    CollegeAdminAssignment,
    CollegeDepartment,
    MentorAssignment,
    MentorChangeRequest,
    SchoolClass,
    SenseiAssignment,
    StaffProfile,
    StudentProfile,
    UserRole,
    Role,
)

User = get_user_model()


# ─── helpers ────────────────────────────────────────────────────────────────

def _json(request: HttpRequest) -> dict:
    try:
        return json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return {}


def _user_mini(user) -> dict:
    profile = getattr(user, "profile", None)
    sp = getattr(user, "student_profile", None)
    sfp = getattr(user, "staff_profile", None)
    pic = None
    if profile and profile.profile_picture:
        try:
            pic = profile.profile_picture.url
        except Exception:
            pass
    return {
        "id": user.id,
        "username": user.username,
        "full_name": user.get_full_name() or user.username,
        "email": user.email,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "profile_picture": pic,
        "register_number": getattr(sp, "register_number", None),
        "staff_id": getattr(sfp, "staff_id", None),
        "college_id": (getattr(sp, "college_id", None) or getattr(sfp, "college_id", None)),
        "department_id": (getattr(sp, "department_id", None) or getattr(sfp, "department_id", None)),
    }


def _get_admin_college(user):
    """Return the college this college_admin manages (first assignment)."""
    assignment = CollegeAdminAssignment.objects.filter(user=user).select_related("college").first()
    return assignment.college if assignment else None


def _has_role(user, *role_names) -> bool:
    if user.is_superuser:
        return True
    return UserRole.objects.filter(user=user, role__name__in=role_names, role__is_active=True).exists()


def _assign_role(user, role_name: str):
    role, _ = Role.objects.get_or_create(name=role_name)
    UserRole.objects.get_or_create(user=user, role=role)


def _generate_password(length=10) -> str:
    chars = string.ascii_letters + string.digits
    return "".join(random.choices(chars, k=length))


# ─── Excel helpers ───────────────────────────────────────────────────────────

def _excel_template(columns: list[str], filename: str) -> HttpResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(columns)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _read_excel_rows(file_obj) -> list[dict]:
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return result


# ════════════════════════════════════════════════════════════════════════════
#  MASTER ADMIN — Colleges
# ════════════════════════════════════════════════════════════════════════════

@require_role("master_admin")
def colleges_list(request: HttpRequest) -> JsonResponse:
    """GET  /api/admin/colleges/          — list / search colleges
    POST /api/admin/colleges/          — create college
    """
    if request.method == "GET":
        q = request.GET.get("q", "").strip()
        qs = College.objects.all().order_by("name")
        if q:
            qs = qs.filter(name__icontains=q)
        data = []
        for c in qs:
            admins = list(
                CollegeAdminAssignment.objects.filter(college=c)
                .select_related("user")
                .values("user__id", "user__username", "user__first_name", "user__last_name", "user__email")
            )
            data.append({
                "id": c.id,
                "name": c.name,
                "code": c.code,
                "city": c.city,
                "admins": [
                    {
                        "id": a["user__id"],
                        "username": a["user__username"],
                        "full_name": f"{a['user__first_name']} {a['user__last_name']}".strip() or a["user__username"],
                        "email": a["user__email"],
                    }
                    for a in admins
                ],
            })
        return JsonResponse({"colleges": data})

    if request.method == "POST":
        body = _json(request)
        name = (body.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name is required"}, status=400)
        college = College.objects.create(
            name=name,
            code=body.get("code", ""),
            city=body.get("city", ""),
        )
        return JsonResponse({"id": college.id, "name": college.name}, status=201)

    return JsonResponse({"detail": "Method not allowed"}, status=405)


@require_role("master_admin")
def college_detail(request: HttpRequest, college_id: int) -> JsonResponse:
    """GET /api/admin/colleges/<id>/  — detail"""
    try:
        c = College.objects.get(pk=college_id)
    except College.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    admins = list(
        CollegeAdminAssignment.objects.filter(college=c)
        .select_related("user")
    )
    return JsonResponse({
        "id": c.id,
        "name": c.name,
        "code": c.code,
        "city": c.city,
        "admins": [_user_mini(a.user) for a in admins],
    })


@require_role("master_admin")
def college_users(request: HttpRequest, college_id: int) -> JsonResponse:
    """GET /api/admin/colleges/<id>/users/?type=student|staff  — users for college"""
    try:
        college = College.objects.get(pk=college_id)
    except College.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    utype = request.GET.get("type", "").strip()
    q = request.GET.get("q", "").strip()

    if utype == "student":
        qs = StudentProfile.objects.filter(college=college).select_related("user", "user__profile")
        if q:
            qs = qs.filter(user__username__icontains=q) | qs.filter(user__first_name__icontains=q) | qs.filter(register_number__icontains=q)
        users = [_user_mini(sp.user) for sp in qs[:100]]
    elif utype == "staff":
        qs = StaffProfile.objects.filter(college=college).select_related("user", "user__profile")
        if q:
            qs = qs.filter(user__username__icontains=q) | qs.filter(user__first_name__icontains=q) | qs.filter(staff_id__icontains=q)
        users = [_user_mini(sfp.user) for sfp in qs[:100]]
    else:
        # Return all users linked to college via student or staff profile
        student_ids = StudentProfile.objects.filter(college=college).values_list("user_id", flat=True)
        staff_ids = StaffProfile.objects.filter(college=college).values_list("user_id", flat=True)
        all_ids = set(student_ids) | set(staff_ids)
        qs = User.objects.filter(id__in=all_ids).select_related("profile", "student_profile", "staff_profile")
        if q:
            qs = qs.filter(username__icontains=q) | qs.filter(first_name__icontains=q)
        users = [_user_mini(u) for u in qs[:100]]

    return JsonResponse({"users": users})


@require_role("master_admin")
def set_college_admin(request: HttpRequest, college_id: int) -> JsonResponse:
    """POST /api/admin/colleges/<id>/set-admin/
    body: { user_id, action: "add"|"remove" }
    """
    if request.method != "POST":
        return JsonResponse({"detail": "POST required"}, status=405)
    try:
        college = College.objects.get(pk=college_id)
    except College.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    body = _json(request)
    user_id = body.get("user_id")
    action = body.get("action", "add")
    if not user_id:
        return JsonResponse({"detail": "user_id required"}, status=400)
    try:
        user = User.objects.get(pk=user_id)
    except User.DoesNotExist:
        return JsonResponse({"detail": "User not found"}, status=404)

    if action == "remove":
        CollegeAdminAssignment.objects.filter(user=user, college=college).delete()
        return JsonResponse({"detail": "Removed"})

    CollegeAdminAssignment.objects.get_or_create(user=user, college=college)
    _assign_role(user, "college_admin")
    return JsonResponse({"detail": "Assigned", "user": _user_mini(user)})


# ════════════════════════════════════════════════════════════════════════════
#  COLLEGE ADMIN — Departments
# ════════════════════════════════════════════════════════════════════════════

@require_role("college_admin")
def departments_list(request: HttpRequest) -> JsonResponse:
    """GET  /api/admin/departments/  — list departments for admin's college
    POST /api/admin/departments/  — create department
    """
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)

    if request.method == "GET":
        depts = CollegeDepartment.objects.filter(college=college).order_by("display_id")
        return JsonResponse({
            "departments": [
                {"id": d.id, "display_id": d.display_id, "dep_name": d.dep_name, "college_id": d.college_id}
                for d in depts
            ]
        })

    if request.method == "POST":
        body = _json(request)
        dep_name = (body.get("dep_name") or "").strip()
        if not dep_name:
            return JsonResponse({"detail": "dep_name required"}, status=400)
        # Auto-assign next display_id for this college
        last = CollegeDepartment.objects.filter(college=college).order_by("-display_id").first()
        display_id = body.get("display_id") or ((last.display_id + 1) if last else 1)
        try:
            dept = CollegeDepartment.objects.create(
                dep_name=dep_name, college=college, display_id=int(display_id)
            )
        except IntegrityError:
            return JsonResponse({"detail": "Department name or display_id already exists"}, status=400)
        return JsonResponse({"id": dept.id, "display_id": dept.display_id, "dep_name": dept.dep_name}, status=201)

    return JsonResponse({"detail": "Method not allowed"}, status=405)


@require_role("college_admin")
def department_detail(request: HttpRequest, dept_id: int) -> JsonResponse:
    """GET/PUT/DELETE /api/admin/departments/<id>/"""
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        dept = CollegeDepartment.objects.get(pk=dept_id, college=college)
    except CollegeDepartment.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        staff = StaffProfile.objects.filter(department=dept).select_related("user", "user__profile")
        students = StudentProfile.objects.filter(department=dept).select_related("user", "user__profile")
        return JsonResponse({
            "id": dept.id,
            "display_id": dept.display_id,
            "dep_name": dept.dep_name,
            "staff": [_user_mini(s.user) for s in staff],
            "students": [_user_mini(s.user) for s in students],
        })

    if request.method == "PUT":
        body = _json(request)
        if "dep_name" in body:
            dept.dep_name = body["dep_name"].strip()
        if "display_id" in body:
            dept.display_id = int(body["display_id"])
        try:
            dept.save()
        except IntegrityError:
            return JsonResponse({"detail": "Duplicate dep_name or display_id"}, status=400)
        return JsonResponse({"id": dept.id, "display_id": dept.display_id, "dep_name": dept.dep_name})

    if request.method == "DELETE":
        dept.delete()
        return JsonResponse({"detail": "Deleted"})

    return JsonResponse({"detail": "Method not allowed"}, status=405)


@require_role("college_admin")
def dept_add_staff(request: HttpRequest, dept_id: int) -> JsonResponse:
    """POST /api/admin/departments/<id>/add-staff/
    body: { user_ids: [id, ...] }  — add existing staff to department
    """
    if request.method != "POST":
        return JsonResponse({"detail": "POST required"}, status=405)
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        dept = CollegeDepartment.objects.get(pk=dept_id, college=college)
    except CollegeDepartment.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    body = _json(request)
    user_ids = body.get("user_ids", [])
    StaffProfile.objects.filter(user_id__in=user_ids, college=college).update(department=dept)
    return JsonResponse({"detail": "Updated"})


@require_role("college_admin")
def dept_add_students(request: HttpRequest, dept_id: int) -> JsonResponse:
    """POST /api/admin/departments/<id>/add-students/
    body: { user_ids: [id, ...] }  — assign existing students to department
    """
    if request.method != "POST":
        return JsonResponse({"detail": "POST required"}, status=405)
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        dept = CollegeDepartment.objects.get(pk=dept_id, college=college)
    except CollegeDepartment.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    body = _json(request)
    user_ids = body.get("user_ids", [])
    StudentProfile.objects.filter(user_id__in=user_ids, college=college).update(department=dept)
    return JsonResponse({"detail": "Updated"})


@require_role("college_admin")
def dept_bulk_upload(request: HttpRequest, dept_id: int) -> JsonResponse | HttpResponse:
    """POST /api/admin/departments/<id>/bulk-upload/
    form-data: type=staff|student, sub_type=existing|new, file=<excel>
    GET with ?template=staff|student to download blank template
    """
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        dept = CollegeDepartment.objects.get(pk=dept_id, college=college)
    except CollegeDepartment.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    # Template download
    if request.method == "GET":
        template = request.GET.get("template", "")
        if template == "student_existing":
            return _excel_template(["register_number"], "template_existing_students.xlsx")
        if template == "student_new":
            return _excel_template(["register_no", "name", "emailid", "department", "password"], "template_new_students.xlsx")
        if template == "staff_existing":
            return _excel_template(["staff_id"], "template_existing_staff.xlsx")
        if template == "staff_new":
            return _excel_template(["staffid", "name", "emailid", "department", "password"], "template_new_staff.xlsx")
        return JsonResponse({"detail": "template param required"}, status=400)

    if request.method != "POST":
        return JsonResponse({"detail": "Method not allowed"}, status=405)

    utype = request.POST.get("type", "student")
    sub_type = request.POST.get("sub_type", "existing")
    file_obj = request.FILES.get("file")
    if not file_obj:
        return JsonResponse({"detail": "file required"}, status=400)

    rows = _read_excel_rows(file_obj)
    created, updated, errors = 0, 0, []

    with transaction.atomic():
        for i, row in enumerate(rows, start=2):
            try:
                if utype == "student":
                    if sub_type == "existing":
                        reg = str(row.get("register_number", "")).strip()
                        if not reg:
                            errors.append(f"Row {i}: missing register_number")
                            continue
                        sp = StudentProfile.objects.filter(register_number=reg, college=college).first()
                        if sp:
                            sp.department = dept
                            sp.save(update_fields=["department"])
                            updated += 1
                        else:
                            errors.append(f"Row {i}: student {reg} not found in college")
                    else:  # new
                        reg = str(row.get("register_no", "")).strip()
                        name = str(row.get("name", "")).strip()
                        email = str(row.get("emailid", "")).strip()
                        dept_id_val = row.get("department")
                        pwd = str(row.get("password", "")).strip() or _generate_password()
                        if not (reg and email):
                            errors.append(f"Row {i}: missing register_no or emailid")
                            continue
                        username = email.split("@")[0]
                        user = User.objects.create(
                            username=username,
                            email=email,
                            first_name=name.split()[0] if name else "",
                            last_name=" ".join(name.split()[1:]) if name else "",
                            password=make_password(pwd),
                        )
                        sp_dept = CollegeDepartment.objects.filter(pk=dept_id_val, college=college).first() if dept_id_val else dept
                        StudentProfile.objects.create(user=user, register_number=reg, college=college, department=sp_dept or dept)
                        _assign_role(user, "student")
                        created += 1
                else:  # staff
                    if sub_type == "existing":
                        sid = str(row.get("staff_id", "")).strip()
                        if not sid:
                            errors.append(f"Row {i}: missing staff_id")
                            continue
                        sfp = StaffProfile.objects.filter(staff_id=sid, college=college).first()
                        if sfp:
                            sfp.department = dept
                            sfp.save(update_fields=["department"])
                            updated += 1
                        else:
                            errors.append(f"Row {i}: staff {sid} not found in college")
                    else:  # new
                        sid = str(row.get("staffid", "")).strip()
                        name = str(row.get("name", "")).strip()
                        email = str(row.get("emailid", "")).strip()
                        dept_id_val = row.get("department")
                        pwd = str(row.get("password", "")).strip() or _generate_password()
                        if not (sid and email):
                            errors.append(f"Row {i}: missing staffid or emailid")
                            continue
                        username = email.split("@")[0]
                        user = User.objects.create(
                            username=username,
                            email=email,
                            first_name=name.split()[0] if name else "",
                            last_name=" ".join(name.split()[1:]) if name else "",
                            password=make_password(pwd),
                        )
                        sfp_dept = CollegeDepartment.objects.filter(pk=dept_id_val, college=college).first() if dept_id_val else dept
                        StaffProfile.objects.create(user=user, staff_id=sid, college=college, department=sfp_dept or dept)
                        _assign_role(user, "staff")
                        created += 1
            except Exception as exc:
                errors.append(f"Row {i}: {exc}")

    return JsonResponse({"created": created, "updated": updated, "errors": errors})


# ════════════════════════════════════════════════════════════════════════════
#  COLLEGE ADMIN — Sensei
# ════════════════════════════════════════════════════════════════════════════

@require_role("college_admin")
def sensei_list(request: HttpRequest) -> JsonResponse:
    """GET  /api/admin/sensei/  — list senseis in admin's college
    POST /api/admin/sensei/  — add a user as sensei  { user_id }
    """
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)

    if request.method == "GET":
        qs = SenseiAssignment.objects.filter(college=college).select_related("user", "user__profile")
        data = []
        for sa in qs:
            classes = SchoolClass.objects.filter(college=college, sensei=sa.user).values("id", "name")
            u = _user_mini(sa.user)
            u["classes"] = list(classes)
            u["assignment_id"] = sa.id
            data.append(u)
        return JsonResponse({"senseis": data})

    if request.method == "POST":
        body = _json(request)
        user_id = body.get("user_id")
        if not user_id:
            return JsonResponse({"detail": "user_id required"}, status=400)
        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return JsonResponse({"detail": "User not found"}, status=404)
        SenseiAssignment.objects.get_or_create(user=user, college=college)
        _assign_role(user, "sensei")
        return JsonResponse({"detail": "Added", "user": _user_mini(user)}, status=201)

    return JsonResponse({"detail": "Method not allowed"}, status=405)


@require_role("college_admin")
def sensei_assign_classes(request: HttpRequest, sensei_user_id: int) -> JsonResponse:
    """GET  /api/admin/sensei/<user_id>/classes/  — get assigned classes
    POST /api/admin/sensei/<user_id>/classes/  — replace assigned classes  { class_ids: [] }
    """
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        sensei_user = User.objects.get(pk=sensei_user_id)
    except User.DoesNotExist:
        return JsonResponse({"detail": "User not found"}, status=404)

    if request.method == "GET":
        classes = SchoolClass.objects.filter(college=college, sensei=sensei_user).values("id", "name")
        return JsonResponse({"classes": list(classes)})

    if request.method == "POST":
        body = _json(request)
        class_ids = body.get("class_ids", [])
        # Remove from previous
        SchoolClass.objects.filter(college=college, sensei=sensei_user).update(sensei=None)
        # Assign
        SchoolClass.objects.filter(id__in=class_ids, college=college).update(sensei=sensei_user)
        return JsonResponse({"detail": "Classes assigned"})

    return JsonResponse({"detail": "Method not allowed"}, status=405)


# ════════════════════════════════════════════════════════════════════════════
#  COLLEGE ADMIN — Classes
# ════════════════════════════════════════════════════════════════════════════

@require_role("college_admin")
def classes_list(request: HttpRequest) -> JsonResponse:
    """GET  /api/admin/classes/  — list classes
    POST /api/admin/classes/  — create class  { name }
    """
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)

    if request.method == "GET":
        qs = SchoolClass.objects.filter(college=college).select_related("sensei")
        return JsonResponse({
            "classes": [
                {
                    "id": c.id,
                    "name": c.name,
                    "sensei_id": c.sensei_id,
                    "sensei_name": c.sensei.get_full_name() if c.sensei else None,
                    "student_count": c.class_students.count(),
                }
                for c in qs
            ]
        })

    if request.method == "POST":
        body = _json(request)
        name = (body.get("name") or "").strip()
        if not name:
            return JsonResponse({"detail": "name required"}, status=400)
        try:
            sc = SchoolClass.objects.create(name=name, college=college)
        except IntegrityError:
            return JsonResponse({"detail": "Class name already exists"}, status=400)
        return JsonResponse({"id": sc.id, "name": sc.name}, status=201)

    return JsonResponse({"detail": "Method not allowed"}, status=405)


@require_role("college_admin")
def class_detail(request: HttpRequest, class_id: int) -> JsonResponse:
    """GET/DELETE /api/admin/classes/<id>/"""
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        sc = SchoolClass.objects.get(pk=class_id, college=college)
    except SchoolClass.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        students = ClassStudent.objects.filter(school_class=sc).select_related("student", "student__profile", "student__student_profile")
        return JsonResponse({
            "id": sc.id,
            "name": sc.name,
            "sensei_id": sc.sensei_id,
            "sensei_name": sc.sensei.get_full_name() if sc.sensei else None,
            "students": [_user_mini(cs.student) for cs in students],
        })

    if request.method == "DELETE":
        sc.delete()
        return JsonResponse({"detail": "Deleted"})

    return JsonResponse({"detail": "Method not allowed"}, status=405)


@require_role("college_admin")
def class_assign_students(request: HttpRequest, class_id: int) -> JsonResponse:
    """POST /api/admin/classes/<id>/assign-students/
    body: { user_ids: [id, ...] }  — add existing students
    """
    if request.method != "POST":
        return JsonResponse({"detail": "POST required"}, status=405)
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        sc = SchoolClass.objects.get(pk=class_id, college=college)
    except SchoolClass.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    body = _json(request)
    user_ids = body.get("user_ids", [])
    added = 0
    for uid in user_ids:
        _, was_created = ClassStudent.objects.get_or_create(school_class=sc, student_id=uid)
        if was_created:
            added += 1
    return JsonResponse({"added": added})


@require_role("college_admin")
def class_bulk_upload(request: HttpRequest, class_id: int) -> JsonResponse | HttpResponse:
    """POST /api/admin/classes/<id>/bulk-upload/
    form-data: sub_type=existing|new, file=<excel>
    GET ?template=existing|new  — download template
    """
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        sc = SchoolClass.objects.get(pk=class_id, college=college)
    except SchoolClass.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        template = request.GET.get("template", "")
        if template == "existing":
            return _excel_template(["register_number"], "template_existing_students.xlsx")
        if template == "new":
            return _excel_template(["register_no", "name", "emailid", "department", "password"], "template_new_students.xlsx")
        return JsonResponse({"detail": "template param required"}, status=400)

    if request.method != "POST":
        return JsonResponse({"detail": "Method not allowed"}, status=405)

    sub_type = request.POST.get("sub_type", "existing")
    file_obj = request.FILES.get("file")
    if not file_obj:
        return JsonResponse({"detail": "file required"}, status=400)

    rows = _read_excel_rows(file_obj)
    created, assigned, errors = 0, 0, []

    with transaction.atomic():
        for i, row in enumerate(rows, start=2):
            try:
                if sub_type == "existing":
                    reg = str(row.get("register_number", "")).strip()
                    if not reg:
                        errors.append(f"Row {i}: missing register_number")
                        continue
                    sp = StudentProfile.objects.filter(register_number=reg, college=college).first()
                    if not sp:
                        errors.append(f"Row {i}: student {reg} not found")
                        continue
                    _, was_created = ClassStudent.objects.get_or_create(school_class=sc, student=sp.user)
                    assigned += 1
                else:  # new
                    reg = str(row.get("register_no", "")).strip()
                    name = str(row.get("name", "")).strip()
                    email = str(row.get("emailid", "")).strip()
                    dept_id_val = row.get("department")
                    pwd = str(row.get("password", "")).strip() or _generate_password()
                    if not (reg and email):
                        errors.append(f"Row {i}: missing register_no or emailid")
                        continue
                    username = email.split("@")[0]
                    user = User.objects.create(
                        username=username, email=email,
                        first_name=name.split()[0] if name else "",
                        last_name=" ".join(name.split()[1:]) if name else "",
                        password=make_password(pwd),
                    )
                    dept_obj = CollegeDepartment.objects.filter(pk=dept_id_val, college=college).first() if dept_id_val else None
                    StudentProfile.objects.create(user=user, register_number=reg, college=college, department=dept_obj)
                    _assign_role(user, "student")
                    ClassStudent.objects.create(school_class=sc, student=user)
                    created += 1
            except Exception as exc:
                errors.append(f"Row {i}: {exc}")

    return JsonResponse({"created": created, "assigned": assigned, "errors": errors})


# ════════════════════════════════════════════════════════════════════════════
#  COLLEGE ADMIN — Students
# ════════════════════════════════════════════════════════════════════════════

@require_role("college_admin")
def students_list(request: HttpRequest) -> JsonResponse:
    """GET /api/admin/students/?dept_id=&class_id=&q="""
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)

    dept_id = request.GET.get("dept_id")
    class_id = request.GET.get("class_id")
    q = request.GET.get("q", "").strip()

    qs = StudentProfile.objects.filter(college=college).select_related("user", "user__profile", "department")

    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if q:
        qs = qs.filter(user__username__icontains=q) | qs.filter(user__first_name__icontains=q) | qs.filter(register_number__icontains=q)

    if class_id:
        student_ids = ClassStudent.objects.filter(school_class_id=class_id).values_list("student_id", flat=True)
        qs = qs.filter(user_id__in=student_ids)

    data = []
    for sp in qs[:200]:
        u = _user_mini(sp.user)
        u["department"] = {"id": sp.department_id, "name": sp.department.dep_name} if sp.department else None
        # class info
        classes = ClassStudent.objects.filter(student=sp.user, school_class__college=college).select_related("school_class")
        u["classes"] = [{"id": cs.school_class.id, "name": cs.school_class.name} for cs in classes]
        # mentor
        ma = MentorAssignment.objects.filter(student=sp.user).select_related("mentor").first()
        u["mentor"] = _user_mini(ma.mentor) if ma else None
        data.append(u)

    return JsonResponse({"students": data})


@require_role("college_admin")
def staff_list(request: HttpRequest) -> JsonResponse:
    """GET /api/admin/staff/?dept_id=&q="""
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)

    dept_id = request.GET.get("dept_id")
    q = request.GET.get("q", "").strip()

    qs = StaffProfile.objects.filter(college=college).select_related("user", "user__profile", "department")
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if q:
        qs = qs.filter(user__username__icontains=q) | qs.filter(user__first_name__icontains=q) | qs.filter(staff_id__icontains=q)

    data = []
    for sfp in qs[:200]:
        u = _user_mini(sfp.user)
        u["department"] = {"id": sfp.department_id, "name": sfp.department.dep_name} if sfp.department else None
        data.append(u)

    return JsonResponse({"staff": data})


# ════════════════════════════════════════════════════════════════════════════
#  COLLEGE ADMIN — Mentor Change Requests
# ════════════════════════════════════════════════════════════════════════════

@require_role("college_admin")
def requests_list(request: HttpRequest) -> JsonResponse:
    """GET  /api/admin/requests/  — list mentor change requests"""
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)

    status_filter = request.GET.get("status", "pending")
    qs = MentorChangeRequest.objects.filter(college=college).select_related(
        "student", "student__profile", "requested_mentor", "requested_mentor__profile"
    ).order_by("-created_at")
    if status_filter != "all":
        qs = qs.filter(status=status_filter)

    data = [
        {
            "id": r.id,
            "student": _user_mini(r.student),
            "requested_mentor": _user_mini(r.requested_mentor) if r.requested_mentor else None,
            "reason": r.reason,
            "status": r.status,
            "created_at": r.created_at.isoformat(),
        }
        for r in qs
    ]
    return JsonResponse({"requests": data})


@require_role("college_admin")
def request_respond(request: HttpRequest, req_id: int) -> JsonResponse:
    """POST /api/admin/requests/<id>/respond/  { action: approve|reject }"""
    if request.method != "POST":
        return JsonResponse({"detail": "POST required"}, status=405)
    college = _get_admin_college(request.user)
    if not college:
        return JsonResponse({"detail": "No college assigned"}, status=403)
    try:
        mcr = MentorChangeRequest.objects.get(pk=req_id, college=college)
    except MentorChangeRequest.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    body = _json(request)
    action = body.get("action", "")
    if action == "approve":
        mcr.status = MentorChangeRequest.Status.APPROVED
        if mcr.requested_mentor:
            MentorAssignment.objects.update_or_create(
                student=mcr.student,
                defaults={"mentor": mcr.requested_mentor},
            )
    elif action == "reject":
        mcr.status = MentorChangeRequest.Status.REJECTED
    else:
        return JsonResponse({"detail": "action must be approve or reject"}, status=400)
    mcr.save()
    return JsonResponse({"detail": f"Request {action}d", "status": mcr.status})


# ════════════════════════════════════════════════════════════════════════════
#  STUDENT — Mentor
# ════════════════════════════════════════════════════════════════════════════

@api_login_required
def student_mentor(request: HttpRequest) -> JsonResponse:
    """GET  /api/student/mentor/   — current mentor
    POST /api/student/mentor/   — set mentor  { mentor_user_id }
    """
    sp = getattr(request.user, "student_profile", None)
    if not sp:
        return JsonResponse({"detail": "Student profile not found"}, status=403)

    if request.method == "GET":
        ma = MentorAssignment.objects.filter(student=request.user).select_related("mentor", "mentor__profile", "mentor__staff_profile").first()
        if not ma:
            # Return available staff from student's department
            dept = sp.department
            college = sp.college
            if dept:
                staff_qs = StaffProfile.objects.filter(department=dept, college=college).select_related("user", "user__profile")
            elif college:
                staff_qs = StaffProfile.objects.filter(college=college).select_related("user", "user__profile")
            else:
                staff_qs = StaffProfile.objects.none()
            return JsonResponse({
                "mentor": None,
                "available_staff": [_user_mini(s.user) for s in staff_qs],
            })
        return JsonResponse({"mentor": _user_mini(ma.mentor), "available_staff": []})

    if request.method == "POST":
        body = _json(request)
        mentor_id = body.get("mentor_user_id")
        if not mentor_id:
            return JsonResponse({"detail": "mentor_user_id required"}, status=400)
        try:
            mentor_user = User.objects.get(pk=mentor_id)
        except User.DoesNotExist:
            return JsonResponse({"detail": "User not found"}, status=404)
        MentorAssignment.objects.update_or_create(
            student=request.user,
            defaults={"mentor": mentor_user},
        )
        return JsonResponse({"detail": "Mentor assigned", "mentor": _user_mini(mentor_user)})

    return JsonResponse({"detail": "Method not allowed"}, status=405)


@api_login_required
def student_mentor_change_request(request: HttpRequest) -> JsonResponse:
    """POST /api/student/mentor/change-request/
    body: { mentor_user_id, reason }
    """
    if request.method != "POST":
        return JsonResponse({"detail": "POST required"}, status=405)
    sp = getattr(request.user, "student_profile", None)
    if not sp or not sp.college:
        return JsonResponse({"detail": "Student profile/college not found"}, status=403)

    body = _json(request)
    mentor_id = body.get("mentor_user_id")
    reason = body.get("reason", "")
    mentor_user = None
    if mentor_id:
        mentor_user = User.objects.filter(pk=mentor_id).first()

    mcr = MentorChangeRequest.objects.create(
        student=request.user,
        requested_mentor=mentor_user,
        college=sp.college,
        reason=reason,
    )
    return JsonResponse({"detail": "Request submitted", "id": mcr.id}, status=201)


# ════════════════════════════════════════════════════════════════════════════
#  STAFF — Mentees
# ════════════════════════════════════════════════════════════════════════════

@require_role("staff", "sensei", "college_admin", "master_admin")
def staff_mentees(request: HttpRequest) -> JsonResponse:
    """GET /api/staff/mentees/"""
    qs = MentorAssignment.objects.filter(mentor=request.user).select_related(
        "student", "student__profile", "student__student_profile", "student__student_profile__department"
    )
    return JsonResponse({
        "mentees": [_user_mini(ma.student) for ma in qs]
    })


# ════════════════════════════════════════════════════════════════════════════
#  USER ROLE INFO
# ════════════════════════════════════════════════════════════════════════════

@api_login_required
def my_roles(request: HttpRequest) -> JsonResponse:
    """GET /api/admin/me/roles/  — return current user's roles.
    Merges administration.UserRole entries with core.User.role so
    ordinary students/staff are included without needing a UserRole row."""
    roles: list[str] = list(
        UserRole.objects.filter(user=request.user, role__is_active=True)
        .values_list("role__name", flat=True)
    )
    # Also include the user-level role field (student / employee→staff)
    profile_role = getattr(request.user, "role", None)
    if profile_role:
        mapped = "staff" if profile_role == "employee" else profile_role
        if mapped not in roles:
            roles.append(mapped)
    if request.user.is_superuser and "master_admin" not in roles:
        roles.append("master_admin")
    # college admin info
    college_admin_of = list(
        CollegeAdminAssignment.objects.filter(user=request.user).select_related("college")
        .values("college__id", "college__name")
    )
    return JsonResponse({
        "roles": roles,
        "college_admin_of": [{"id": c["college__id"], "name": c["college__name"]} for c in college_admin_of],
    })
